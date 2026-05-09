#!/usr/bin/env python3
"""
Build docs/commercial-workbook/Commercial_Readiness_Workbook.xlsx from CSV sources.

  pip install -r docs/commercial-workbook/requirements-workbook.txt
  python3 scripts/build_commercial_workbook.py

Requires: openpyxl
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
WB_DIR = ROOT / "docs" / "commercial-workbook"
OUT_PATH = WB_DIR / "Commercial_Readiness_Workbook.xlsx"


def _a(key: str) -> str:
    """Excel formula: lookup numeric assumption by stable key (column A = keys, B = values)."""
    return (
        f'INDEX(Assumptions!$B:$B,MATCH("{key}",Assumptions!$A:$A,0))'
    )


def build_assumptions(ws) -> None:
    path = WB_DIR / "assumptions.csv"
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)


def append_csv_sheet(wb, name: str, csv_name: str) -> None:
    ws = wb.create_sheet(title=name)
    path = WB_DIR / csv_name
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)


def build_monthly_pl(ws) -> None:
    """Monthly P&L with formulas wired to Assumptions (Excel INDEX/MATCH)."""
    ws.append(["section", "line_item", "amount_usd", "driver_or_formula_notes"])

    # Revenue (rows 2-4 in 1-based = after header row 1, data rows start 2)
    ws.append(["Revenue", "Subscription MRR", f"={_a('tenant_count')}*{_a('revenue_per_tenant_month')}", "tenant_count * revenue_per_tenant_month"])
    ws.append(["Revenue", "Professional services (one-time)", 0, "Edit if booked"])
    ws.append(["Revenue", "Other", 0, "Optional"])

    ws.append([])  # blank row

    ws.append(
        [
            "COGS",
            "LLM API (if you absorb)",
            f"={_a('tenant_count')}*{_a('jobs_per_tenant_per_day')}*{_a('days_per_month')}*"
            f"(({_a('tokens_in_per_job')}/1000000)*{_a('llm_price_in_per_million_tokens')}+"
            f"({_a('tokens_out_per_job')}/1000000)*{_a('llm_price_out_per_million_tokens')})",
            "Fleet token cost",
        ]
    )
    ws.append(
        [
            "COGS",
            "Object storage",
            f"={_a('tenant_count')}*{_a('storage_gb_per_tenant_month')}*{_a('storage_usd_per_gb_month')}",
        ]
    )
    ws.append(
        [
            "COGS",
            "Egress",
            f"={_a('tenant_count')}*{_a('egress_gb_per_tenant_month')}*{_a('egress_usd_per_gb')}",
        ]
    )
    ws.append(
        [
            "COGS",
            "Compute (shared pool)",
            f"={_a('shared_vcpu_for_app_pool')}*24*{_a('days_per_month')}*{_a('vcpu_hour_usd')}",
        ]
    )
    ws.append(["COGS", "Managed database", f"={_a('managed_db_usd_month')}"])
    ws.append(["COGS", "Fixed vendor overhead (COGS allocation)", f"={_a('fixed_saas_overhead_usd_month')}"])
    ws.append(
        [
            "COGS",
            "Payment processing fees",
            f"={_a('tenant_count')}*({_a('stripe_fee_percent')}*{_a('revenue_per_tenant_month')}+{_a('stripe_fixed_usd_per_charge')})",
        ]
    )

    ws.append([])

    # Row indices (1-based): header=1, revenue 2-4, blank 5, COGS 6-12, blank 13
    first_cogs = 6
    last_cogs = 12
    ws.append(
        [
            "Gross profit",
            "Gross profit (after COGS incl. payment fees)",
            f"=SUM(C2:C4)-SUM(C{first_cogs}:C{last_cogs})",
        ]
    )
    gross_row = 14

    ws.append([])

    # Row 15 is blank; OpEx block starts row 16
    payroll_row = 16
    ws.append(
        [
            "OpEx",
            "Payroll (fully loaded)",
            f"=({_a('headcount_fte_sales')}+{_a('headcount_fte_engineering')}+{_a('headcount_fte_support')})*{_a('fully_loaded_cost_usd_per_fte_month')}",
        ]
    )
    ws.append(["OpEx", "Legal", f"={_a('legal_retainer_usd_month')}"])
    ws.append(["OpEx", "Accounting", f"={_a('accounting_bookkeeping_usd_month')}"])
    ws.append(["OpEx", "Insurance", f"={_a('insurance_usd_month')}"])
    ws.append(["OpEx", "Marketing and demand gen", 0, "Edit as you spend"])
    ws.append(["OpEx", "Travel and customer visits", 0])
    ws.append(["OpEx", "Software and tooling (non-COGS)", 0])

    ws.append([])

    first_opex = payroll_row
    last_opex = payroll_row + 6
    ws.append(
        [
            "EBITDA",
            "EBITDA (simplified)",
            f"=C{gross_row}-SUM(C{first_opex}:C{last_opex})",
        ]
    )

    ws.append(
        [
            "Cash",
            "Illustrative: cash impact vs EBITDA",
            "See runway model; link founder_draw and fundraising separately",
            "Not a full cash flow statement",
        ]
    )


def build_unit_economics(ws) -> None:
    ws.append(["metric", "value_usd_or_count", "notes"])
    ws.append(["Jobs per tenant per month", f"={_a('jobs_per_tenant_per_day')}*{_a('days_per_month')}"])
    ws.append(
        [
            "Token cost per job",
            f"=({_a('tokens_in_per_job')}/1000000)*{_a('llm_price_in_per_million_tokens')}+({_a('tokens_out_per_job')}/1000000)*{_a('llm_price_out_per_million_tokens')}",
        ]
    )
    ws.append(
        [
            "LLM COGS per tenant per month",
            f"=C2*C3",
            "jobs/month * cost/job",
        ]
    )
    ws.append(
        [
            "Fleet LLM COGS per month",
            f"={_a('tenant_count')}*C4",
        ]
    )
    ws.append(
        [
            "Allocated compute per tenant (shared pool split)",
            f"=({_a('shared_vcpu_for_app_pool')}*24*{_a('days_per_month')}*{_a('vcpu_hour_usd')})/{_a('tenant_count')}",
        ]
    )
    ws.append(
        [
            "Storage + egress per tenant per month",
            f"={_a('storage_gb_per_tenant_month')}*{_a('storage_usd_per_gb_month')}+{_a('egress_gb_per_tenant_month')}*{_a('egress_usd_per_gb')}",
        ]
    )
    ws.append(
        [
            "Variable COGS per tenant (LLM + storage + egress + allocated compute)",
            f"=C4+C6+C7",
        ]
    )
    ws.append(
        [
            "MRR per tenant",
            f"={_a('revenue_per_tenant_month')}",
        ]
    )
    ws.append(
        [
            "Contribution margin per tenant (MRR - variable COGS; excludes DB/overhead allocation)",
            f"=C9-C8",
        ]
    )
    ws.append(
        [
            "Simple LTV (revenue / monthly churn rate) — illustrative only",
            f"=C9/(({_a('annual_churn_percent')}/100)/12)",
        ]
    )
    ws.append(
        [
            "CAC payback months (CAC / monthly contribution)",
            f"={_a('customer_acquisition_cost_usd')}/C10",
        ]
    )


def main() -> int:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Missing openpyxl. Install with:", file=sys.stderr)
        print("  pip install -r docs/commercial-workbook/requirements-workbook.txt", file=sys.stderr)
        return 1

    wb = Workbook()

    # Assumptions first (referenced by formulas)
    ws_a = wb.active
    ws_a.title = "Assumptions"
    build_assumptions(ws_a)

    ws_pl = wb.create_sheet("Monthly_P_L")
    build_monthly_pl(ws_pl)

    ws_ue = wb.create_sheet("Unit_Economics")
    build_unit_economics(ws_ue)

    append_csv_sheet(wb, "Pricing", "pricing_packaging.csv")
    append_csv_sheet(wb, "GTM_ICP", "gtm_icp.csv")
    append_csv_sheet(wb, "Readiness", "readiness_checklist.csv")
    append_csv_sheet(wb, "Risks", "risks_mitigations.csv")
    append_csv_sheet(wb, "KPIs", "kpi_dashboard.csv")

    # Widen first column on formula sheets for readability
    for sheet_name in ("Monthly_P_L", "Unit_Economics"):
        ws = wb[sheet_name]
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 50

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
